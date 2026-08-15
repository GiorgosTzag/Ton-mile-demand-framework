"""
build_outputs.py — regenerate the live xlsx model and HTML one-pager (2025 baseline).
Two levers: Red Sea rerouting (distance, continuous 0-100%) and net Gulf export loss
(volume, 0-100%). Excel formulas are flag-driven and include a Baseline vs Scenario
output block. Run: python build_outputs.py
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import tonmile_engine as E

OUT = Path(__file__).parent
BLUE = Font(name="Arial", color="0000FF", size=10)
BLACK = Font(name="Arial", color="000000", size=10)
FLAG = Font(name="Arial", color="0E7C86", size=10, bold=True)
BOLD = Font(name="Arial", bold=True, size=10)
TITLE = Font(name="Arial", bold=True, size=14)
HEADER_FILL = PatternFill("solid", fgColor="0B1F2A")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
YELLOW = PatternFill("solid", fgColor="FFFF00")
OUT_FILL = PatternFill("solid", fgColor="EEF3F5")
thin = Side(style="thin", color="B8C4CC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

lanes = E.load_lanes()


def build_xlsx():
    wb = Workbook(); ws = wb.active; ws.title = "Ton-Mile Model"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Tanker Ton-Mile Demand — 2025 Baseline Scenario Model"; ws["A1"].font = TITLE
    ws["A2"] = ("Ton-miles = cargo x distance. Selected major lanes (~72% of Hormuz oil flows), not the whole "
                "market. Two physical levers; every row runs the same flag-driven formula.")
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="555555")

    # ---- levers ----
    ws["A4"] = "SCENARIO LEVERS"; ws["A4"].font = BOLD
    ws["A5"] = "Red Sea rerouting  (share of exposed vessels diverting via Cape, 0–100%)"
    ws["A6"] = "Net Gulf export loss  (Hormuz: share of Gulf-origin cargo lost, 0–100%)"
    for r in (5, 6):
        c = ws[f"H{r}"]; c.value = 0; c.font = BLUE; c.fill = YELLOW; c.border = BORDER; c.number_format = "0%"
    ws["A7"] = "↑ Yellow = inputs (type e.g. 50%). Red Sea drives distance; Gulf loss drives volume."
    ws["A7"].font = Font(name="Arial", italic=True, size=9, color="C7632E")

    # ---- scenario output block (filled after table refs resolve; uses forward refs) ----
    hdr = 15
    first = hdr + 1
    last = first + len(lanes) - 1
    ws["A9"] = "SCENARIO OUTPUT  (bn tonne-miles)"; ws["A9"].font = BOLD
    for lbl, rr, formula in [
        ("Baseline (levers off)", 10, f"=SUM(P{first}:P{last})"),
        ("Scenario (as set)", 11, f"=SUM(O{first}:O{last})"),
        ("Δ ton-miles", 12, "=C11-C10"),
        ("Δ %", 13, "=(C11-C10)/C10"),
    ]:
        ws.cell(rr, 1, lbl).font = BLACK
        c = ws.cell(rr, 3, formula); c.font = BOLD if rr in (10, 11) else BLACK
        c.fill = OUT_FILL; c.border = BORDER
        c.number_format = "+0.0%;-0.0%" if rr == 13 else "#,##0"
    ws.column_dimensions["C"].width = 13

    # ---- lane table ----
    headers = ["ID", "Segment", "Trade", "Origin", "Destination", "Volume\n(mb/d)",
               "Dist base\n(nm)", "Dist Cape\n(nm)", "RedSea", "Hormuz", "bbl/t",
               "Eff. vol\n(mb/d)", "Eff. dist\n(nm)", "Tonnes/yr\n(m)", "Scenario\nTM (bn)", "Baseline\nTM (bn)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hdr, c, h); cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = BORDER
    ws.row_dimensions[hdr].height = 30

    for i, L in enumerate(lanes):
        r = first + i
        for c, v in enumerate([int(L["id"]), L["segment"], L["trade"], L["origin"], L["destination"]], 1):
            ws.cell(r, c, v).font = BLACK
        ws.cell(r, 6, L["volume_mbd"]).font = BLUE
        ws.cell(r, 7, L["dist_base_nm"]).font = BLUE
        ws.cell(r, 8, L["dist_cape_nm"]).font = BLUE
        ws.cell(r, 9, "Y" if L["red_sea"] else "N").font = FLAG       # I  RedSea flag
        ws.cell(r, 10, "Y" if L["hormuz"] else "N").font = FLAG       # J  Hormuz flag
        ws.cell(r, 11, L["bbl_per_t"]).font = BLUE                    # K  bbl/t
        # eff vol: Gulf export loss on Hormuz-flagged lanes
        ws.cell(r, 12, f'=MAX(0,F{r}*IF(J{r}="Y",1-$H$6,1))').font = BLACK
        # eff dist: CONTINUOUS Red Sea reroute -> base + share*(cape-base), on RedSea-flagged lanes
        ws.cell(r, 13, f'=G{r}+IF(I{r}="Y",$H$5,0)*(H{r}-G{r})').font = BLACK
        ws.cell(r, 14, f"=L{r}*1000000*365/K{r}/1000000").font = BLACK          # tonnes/yr (m)
        ws.cell(r, 15, f"=N{r}*1000000*M{r}/1000000000").font = BLACK           # scenario TM (bn)
        ws.cell(r, 16, f"=F{r}*1000000*365/K{r}*G{r}/1000000000").font = BLACK   # baseline TM (bn)
        for c in range(1, 17):
            ws.cell(r, c).border = BORDER
        for c in (6, 12, 15, 16):
            ws.cell(r, c).number_format = "#,##0.0"
        for c in (7, 8, 13, 14):
            ws.cell(r, c).number_format = "#,##0"
        ws.cell(r, 11).number_format = "0.00"
        for c in (9, 10):
            ws.cell(r, c).alignment = Alignment(horizontal="center")

    # ---- summary ----
    s = last + 2
    ws.cell(s, 1, "SUMMARY").font = BOLD
    ws.cell(s, 6, "Ton-miles (bn)").font = BOLD
    ws.cell(s, 7, "Share of modelled").font = BOLD
    ws.cell(s + 1, 1, "TOTAL (scenario)").font = BOLD
    ws.cell(s + 1, 6, f"=SUM(O{first}:O{last})").font = BOLD
    ws.cell(s + 1, 6).number_format = "#,##0"
    for k, seg in enumerate(["VLCC", "Suezmax", "Aframax", "LR2", "LR1", "MR"]):
        rr = s + 2 + k
        ws.cell(rr, 1, seg).font = BLACK
        ws.cell(rr, 6, f'=SUMIF(B{first}:B{last},A{rr},O{first}:O{last})').font = BLACK
        ws.cell(rr, 6).number_format = "#,##0"
        ws.cell(rr, 7, f"=F{rr}/$F${s+1}").font = BLACK
        ws.cell(rr, 7).number_format = "0.0%"
    cd = s + 2 + 6
    for k, (lbl, key) in enumerate([("Crude (dirty)", "dirty"), ("Clean (products)", "clean")]):
        rr = cd + k
        ws.cell(rr, 1, lbl).font = BLACK
        ws.cell(rr, 6, f'=SUMIF(C{first}:C{last},"{key}",O{first}:O{last})').font = BLACK
        ws.cell(rr, 6).number_format = "#,##0"
        ws.cell(rr, 7, f"=F{rr}/$F${s+1}").font = BLACK
        ws.cell(rr, 7).number_format = "0.0%"
    ws.cell(cd + 2, 1, "Shares are of MODELLED ton-miles (this 25-lane sample), NOT the global tanker market. "
                       "Sample skews to crude (captures ~80% of Hormuz crude vs ~half of Hormuz products).").font = \
        Font(name="Arial", italic=True, size=8, color="C7632E")
    ws.cell(cd + 3, 1, "Source: illustrative major-lane volumes (Energy Institute / IEA / EIA order of magnitude, "
                       "2025 baseline); one-way laden distances approximate. Not freight rates — demand (cargo-work) only.").font = \
        Font(name="Arial", italic=True, size=8, color="777777")

    for c, w in zip(range(1, 17), [5, 10, 7, 17, 16, 8, 9, 9, 7, 7, 6, 9, 9, 10, 9, 9]):
        ws.column_dimensions[chr(64 + c)].width = w

    chart = BarChart(); chart.type = "bar"; chart.title = "Scenario ton-mile demand by lane (bn)"
    chart.height = 12; chart.width = 19
    chart.add_data(Reference(ws, min_col=15, min_row=hdr, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=5, min_row=first, max_row=last))
    chart.legend = None
    ws.add_chart(chart, f"R{hdr}")
    p = OUT / "tonmile_model.xlsx"; wb.save(p); return p


def build_html():
    rows = E.compute(lanes); total, by_seg, by_trade = E.summarize(rows)
    rs, _, _ = E.summarize(E.compute(lanes, redsea=1.0))
    clo, _, _ = E.summarize(E.compute(lanes, hormuz_loss=0.5))
    rs_d = (rs / total - 1) * 100; clo_d = (clo / total - 1) * 100
    top = sorted(rows, key=lambda r: -r["tm_bn"])[:12]; maxtm = top[0]["tm_bn"]
    seg_order = sorted(by_seg.items(), key=lambda x: -x[1])
    crude_pct = by_trade["dirty"] / total * 100; clean_pct = by_trade["clean"] / total * 100
    AMBER, TEAL, INK = "#C7632E", "#0E7C86", "#0B1F2A"

    def bar(r):
        w = r["tm_bn"] / maxtm * 100
        col = AMBER if r["trade"] == "dirty" else TEAL
        return f"""<div class="lane"><div class="lane-label">{r['origin']} → {r['destination']}</div>
        <div class="lane-track"><div class="lane-bar" style="width:{w:.1f}%;background:{col}"></div>
        <span class="lane-val">{r['tm_bn']:,.0f}</span></div>
        <div class="lane-meta">{r['segment']} · {r['eff_volume']:.1f} mb/d · {r['eff_distance']:,.0f} nm</div></div>"""

    seg_bars = "".join(
        f'<div class="seg"><span>{s}</span><div class="seg-track"><div class="seg-bar" '
        f'style="width:{v/total*100:.1f}%"></div></div><b>{v/total*100:.1f}%</b></div>'
        for s, v in seg_order)
    lanes_html = "".join(bar(r) for r in top)

    def lever(delta, colr, title, body):
        return (f'<div class="lever"><div class="d" style="color:{colr}">{delta:+.1f}%</div>'
                f'<div class="lt">{title}</div><div class="t">{body}</div></div>')

    levers = (
        lever(rs_d, AMBER, "Red Sea rerouting · distance",
              "Full diversion of Suez-transiting lanes via the Cape. Same cargo, longer voyage — <b>more</b> ton-miles. "
              "(Lever is continuous: model any share of vessels diverting.)") +
        lever(clo_d, TEAL, "Net Gulf export loss · volume",
              "Lose 50% of Gulf-origin cargo. Demand <b>falls</b> — the 2026 rate spike came from supply disruption "
              "&amp; risk premia, not this channel."))

    html = f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>Tanker Ton-Mile Map — 2025 Baseline</title>
<style>
:root{{--ink:{INK};--amber:{AMBER};--teal:{TEAL};--paper:#E8EEF1;--line:#C2D0D8;}}
*{{box-sizing:border-box}}
body{{margin:0;background:var(--paper);color:var(--ink);font-family:ui-sans-serif,"Helvetica Neue",Arial,sans-serif;
 background-image:linear-gradient(var(--line) 1px,transparent 1px),linear-gradient(90deg,var(--line) 1px,transparent 1px);
 background-size:32px 32px;background-position:-1px -1px}}
.wrap{{max-width:1040px;margin:0 auto;padding:40px 28px 56px}}
.eyebrow{{font-family:ui-monospace,"SF Mono",Menlo,Consolas,monospace;font-size:11px;letter-spacing:.22em;
 text-transform:uppercase;color:var(--teal);margin:0 0 6px}}
h1{{font-family:"Arial Narrow",Arial,sans-serif;font-weight:800;letter-spacing:-.01em;text-transform:uppercase;
 font-size:clamp(30px,5vw,52px);line-height:.96;margin:0 0 8px}}
.sub{{max-width:64ch;font-size:15px;line-height:1.5;color:#33474f;margin:0 0 28px}}
.headline{{display:flex;flex-wrap:wrap;gap:14px;margin:0 0 30px}}
.stat{{background:var(--ink);color:#fff;padding:16px 18px;flex:1;min-width:150px}}
.stat .n{{font-family:ui-monospace,Menlo,monospace;font-size:28px;font-weight:700;line-height:1}}
.stat .l{{font-size:11px;letter-spacing:.08em;text-transform:uppercase;color:#9fb4bd;margin-top:8px}}
.stat.amber{{background:var(--amber)}} .stat.teal{{background:var(--teal)}}
.grid{{display:grid;grid-template-columns:1.5fr 1fr;gap:34px}}
@media(max-width:820px){{.grid{{grid-template-columns:1fr}}}}
.panel-h{{font-family:ui-monospace,Menlo,monospace;font-size:11px;letter-spacing:.18em;text-transform:uppercase;
 color:#5a6e76;border-bottom:2px solid var(--ink);padding-bottom:6px;margin:0 0 6px}}
.panel-note{{font-size:10.5px;color:#7a8b92;margin:0 0 12px}}
.lane{{margin-bottom:12px}} .lane-label{{font-size:13px;font-weight:600;margin-bottom:3px}}
.lane-track{{position:relative;background:#d4dee4;height:19px}}
.lane-bar{{height:100%;transition:width .9s cubic-bezier(.2,.8,.2,1)}}
.lane-val{{position:absolute;right:6px;top:50%;transform:translateY(-50%);font-family:ui-monospace,Menlo,monospace;
 font-size:11px;font-weight:700;color:var(--ink)}}
.lane-meta{{font-family:ui-monospace,Menlo,monospace;font-size:10.5px;color:#65797f;margin-top:2px}}
.seg{{display:flex;align-items:center;gap:10px;margin-bottom:10px;font-size:13px}}
.seg span{{width:62px;font-weight:600}} .seg b{{width:44px;text-align:right;font-family:ui-monospace,Menlo,monospace}}
.seg-track{{flex:1;background:#d4dee4;height:14px}} .seg-bar{{height:100%;background:var(--ink)}}
.split{{display:flex;height:30px;margin:6px 0 6px;font-family:ui-monospace,Menlo,monospace;font-size:12px;color:#fff;font-weight:700}}
.split .c{{background:var(--amber);display:flex;align-items:center;padding-left:12px}}
.split .p{{background:var(--teal);display:flex;align-items:center;padding-left:12px}}
.lever{{background:#fff;border:1px solid var(--line);padding:14px 16px;margin-bottom:12px}}
.lever .d{{font-family:ui-monospace,Menlo,monospace;font-size:26px;font-weight:700}}
.lever .lt{{font-size:11px;font-family:ui-monospace,Menlo,monospace;letter-spacing:.06em;text-transform:uppercase;color:#5a6e76;margin:2px 0 6px}}
.lever .t{{font-size:13px;line-height:1.45;color:#33474f}}
.legend{{display:flex;gap:18px;font-size:12px;margin:8px 0 0}}
.legend i{{display:inline-block;width:12px;height:12px;margin-right:6px;vertical-align:-1px}}
.foot{{margin-top:30px;font-size:10.5px;color:#7a8b92;border-top:1px solid var(--line);padding-top:12px;line-height:1.5}}
.key{{font-weight:800;color:var(--ink)}}
</style></head><body><div class="wrap">
<p class="eyebrow">Tanker demand · 2025 baseline</p>
<h1>The Ton-Mile Map</h1>
<p class="sub">Tanker demand is not barrels — it is <span class="key">barrels × distance</span>.
Two chokepoints, two different mechanics: the <span class="key">Red Sea</span> is a transit you can reroute around (distance);
<span class="key">Hormuz</span> is the only exit from the Gulf, so lost cargo removes barrels (volume).</p>
<div class="headline">
  <div class="stat"><div class="n">{total:,.0f}</div><div class="l">bn tonne-nm/yr · modelled lanes</div></div>
  <div class="stat amber"><div class="n">{crude_pct:.0f}%</div><div class="l">Crude / dirty*</div></div>
  <div class="stat teal"><div class="n">{clean_pct:.0f}%</div><div class="l">Clean / products*</div></div>
</div>
<div class="grid">
  <div>
    <p class="panel-h">Top lanes by ton-mile demand</p>
    {lanes_html}
    <div class="legend"><span><i style="background:{AMBER}"></i>Crude</span><span><i style="background:{TEAL}"></i>Products</span></div>
  </div>
  <div>
    <p class="panel-h">Share by segment</p>
    <p class="panel-note">of modelled ton-miles — not the global market</p>
    {seg_bars}
    <p class="panel-h" style="margin-top:22px">Crude vs clean</p>
    <div class="split"><div class="c" style="width:{crude_pct:.1f}%">CRUDE {crude_pct:.0f}%</div><div class="p" style="width:{clean_pct:.1f}%">CLEAN</div></div>
    <p class="panel-note">*sample skews to crude (captures ~80% of Hormuz crude vs ~half of products)</p>
    <p class="panel-h" style="margin-top:14px">Two levers, two mechanisms</p>
    {levers}
  </div>
</div>
<p class="foot"><span class="key">Method:</span> ton-miles = cargo (tonnes/yr) × one-way laden distance (nm).
Red Sea rerouting affects Suez-transiting (Europe-bound) lanes only, continuously (0–100% of exposed vessels);
net Gulf export loss affects all Gulf-origin lanes. Selected major lanes (~72% of Hormuz oil flows), 2025 baseline —
volumes illustrative (Energy Institute / IEA / EIA order of magnitude); distances approximate.
One source of truth: trade_lanes.csv → engine → this map + the live Excel model.
This is the demand (cargo-work) half of the rate, not freight rates. Refresh & lock to one base year before interview use.</p>
</div></body></html>"""
    p = OUT / "tonmile_map.html"; p.write_text(html); return p


if __name__ == "__main__":
    print("built:", build_xlsx().name, build_html().name)
